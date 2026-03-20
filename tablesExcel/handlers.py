from openpyxl import Workbook
from openpyxl.styles import Side, Border, Alignment, Font, PatternFill
from openpyxl.utils import get_column_interval, get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from database.models import Month
from .meta import *
from .utils import global_print_setting, check_path_new_file


class PersonalTableRender:
    def __init__(self):
        self._book: Workbook = Workbook()
        self._worksheet: Worksheet = self._book.active
        self._worksheet.title = '1'
        self.prefill_worksheet()

    def save_file(self, file_name='test'):
        try:
            name = f'{file_name}.xlsx'
            self._book.save(name)
        except PermissionError:
            name = check_path_new_file(f'{file_name}.xlsx')
            self._book.save(name)
        return name

    def add_worksheet(self):
        self._worksheet = self._book.create_sheet(f'{len(self._book.sheetnames) + 1}')
        self.prefill_worksheet()

    def fill_data(self, worker, month, calc, target, data_task, data_old_task):
        offset = 0
        for num in range(1, month.days + 1):
            if num < 16:
                cell = self._worksheet.cell(7, 13 + num)
            else:
                cell = self._worksheet.cell(8, num - 2)
            cell.value = num
            cell.font = Font(size='9')

        self._worksheet.cell(1, 20).value = f'{month.year} года'
        self._worksheet.cell(1, 18).value = str(month).lower()
        self._worksheet.cell(4, 28).value = target
        self._worksheet.cell(4, 4).value = worker.get_short_name()
        self._worksheet.cell(4, 12).value = worker.table_num
        self._worksheet.cell(4, 17).value = worker.function.post
        self._worksheet.cell(22, 13).value = ProgramSetting.get_setting().resp_name
        self._worksheet.cell(25, 13).value = ProgramSetting.get_setting().head_name
        self._worksheet.cell(28, 13).value = worker.get_short_name()
        self._worksheet.cell(21, 30).value = sum(calc.get('first_half', 0))
        self._worksheet.cell(23, 30).value = sum(calc.get('second_half', 0))

        self._worksheet.cell(4, 4).alignment = Alignment(horizontal='center', vertical='center')
        self._worksheet.cell(4, 12).alignment = Alignment(horizontal='center', vertical='center')
        self._worksheet.cell(4, 17).alignment = Alignment(horizontal='center', vertical='center')
        self._worksheet.cell(4, 28).alignment = Alignment(horizontal='center', vertical='center')
        for i, task in enumerate(data_task):
            self._worksheet.cell(9 + offset + i, 1).value = i + 1
            data_string = f'{task.is_type} '
            data_string += f'{task.order}\n' if task.order else '\n'
            data_string += f'{task.order.type_obj} {task.order.title} {task.order.name}\n' if task.order else ''
            data_string += f'{task.order.article}' if task.order else f'{task.comment}'
            self._worksheet.cell(9 + offset + i, 2).value = data_string
            self._worksheet.cell(9 + offset + i, 2).alignment += Alignment(wrapText=True)
            self._worksheet.cell(9 + offset + i, 12).value = task.deadline - sum(data_old_task[i].time_worked)
            self._worksheet.cell(9 + offset + i, 32).value = sum(data_old_task[i].time_worked)
            self._worksheet.cell(9 + offset + i, 30).value = sum(task.time_worked)
            for period in task.time_worked:
                if period.date.day <= 15:
                    self._worksheet.cell(9 + offset + i, 13 + period.date.day).value = period.value
                else:
                    self._worksheet.cell(9 + offset + i + 1, period.date.day - 2).value = period.value
            offset += 1

    def prefill_worksheet(self):

        # задать высоту строк
        row_height = [15 for _ in range(5)]
        row_height += [18, 16.5, 16.5]
        row_height += [29.5 for _ in range(12)]
        row_height += [15 for _ in range(10)]
        for key, val in enumerate(row_height):
            # self._worksheet.row_dimensions[key + 6].height = val  # реальное значение
            self._worksheet.row_dimensions[key + 1].height = val  # реальное значение

        # задать ширину ячеек
        for idx in get_column_interval(2, 31):
            self._worksheet.column_dimensions[idx].width = 5  # значение в пикселях умноженное на 7
        self._worksheet.column_dimensions['A'].width = 6

        # объединение ячеек
        lst_range_cell = ['A6:A8', 'B6:K8', 'L6:M8', 'N6:AC6', 'AD6:AE8']
        lst_range_cell += ['A9:A10', 'B9:K10', 'L9:M10', 'AD9:AE10']
        lst_range_cell += ['A11:A12', 'B11:K12', 'L11:M12', 'AD11:AE12']
        lst_range_cell += ['A13:A14', 'B13:K14', 'L13:M14', 'AD13:AE14']
        lst_range_cell += ['A15:A16', 'B15:K16', 'L15:M16', 'AD15:AE16']
        lst_range_cell += ['A17:A18', 'B17:K18', 'L17:M18', 'AD17:AE18']
        lst_range_cell += ['A19:A20', 'B19:K20', 'L19:M20', 'AD19:AE20']
        lst_range_cell += ['W21:AC22', 'W23:AC24', 'AD21:AE22', 'AD23:AE24', 'AD25:AE26', 'AB25:AC26']
        lst_range_cell += ['D4:G4', 'l4:M4', 'Q4:U4', 'AB4:AC4', 'R1:S1']
        [self._worksheet.merge_cells(range_string) for range_string in lst_range_cell]

        # границы и выравнивание в ячейках таблицы
        table = self._worksheet['A6:AE20'] + self._worksheet['AD21:AE26']
        thins = Side(border_style="thin", color="000000")
        for line in table:
            for cell in line:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)

        # нижнее подчеркивание
        underline = (self._worksheet['D4:G4'] + self._worksheet['L4:M4'] + self._worksheet['Q4:U4'] +
                     self._worksheet['R1:S1'] + self._worksheet['H22:M22'] + self._worksheet['H25:M25'] +
                     self._worksheet['H28:M28'] + self._worksheet['AB4:AC4'])
        for line in underline:
            for cell in line:
                cell.border = Border(bottom=thins)

        # предзаполнение текста заголовка и подписей экселя
        for i, adr in enumerate(['Q1', 'K4', 'P4', 'AA4', 'G22', 'M22', 'G25', 'M25', 'M28']):
            cell = self._worksheet[adr]
            if adr in ['M22', 'M25', 'M28']:
                cell.font = Font(size='11')
            else:
                cell.font = Font(size='14')
            if adr != 'M28':
                cell.value = TEMPLATE_ALIGNMENT_RIGHT[i]
            cell.alignment = Alignment(horizontal='right', vertical='center')
        for adr in ['C4', 'G28']:
            cell = self._worksheet[adr]
            cell.font = Font(size='14')
            cell.value = TEMPLATE_ALIGNMENT_RIGHT[-1]
            cell.alignment = Alignment(horizontal='right', vertical='center')
        self._worksheet['T1'].alignment = Alignment(horizontal='left', vertical='center')
        self._worksheet['T1'].font = Font(size='14')

        # заполнение надстрочных пояснений
        for adr in ['D5', 'S5', 'H23', 'H26', 'H29']:
            self._worksheet[adr].font = Font(vertAlign='superscript', size='14')
            if adr != 'S5':
                self._worksheet[adr].alignment = Alignment(horizontal='left', vertical='center')
                if adr == 'D5':
                    self._worksheet[adr].value = TEMPLATE_CLARIFICATION[0]
                else:
                    self._worksheet[adr].value = TEMPLATE_CLARIFICATION[-1]
            else:
                self._worksheet[adr].alignment = Alignment(horizontal='center', vertical='center')
                self._worksheet[adr].value = TEMPLATE_CLARIFICATION[1]

        # предзаполнение таблицы
        for i, adr in enumerate(['A6', 'B6', 'L6', 'N6', 'AD6', 'W21', 'W23', 'AB25']):
            self._worksheet[adr].value = TEMPLATE_TABLE[i]
            self._worksheet[adr].font = Font(size='11')
            self._worksheet[adr].alignment += Alignment(wrapText=True)
            if adr in ['W21', 'W23', 'AB25']:
                self._worksheet[adr].alignment += Alignment(horizontal='right', vertical='center')
        self._worksheet['AD25'] = '=SUM(AD21:AE24)'

        # заливка ячеек таблицы
        range_cell = []
        for adr in range(8, 21, 2):
            range_cell += self._worksheet[f'N{adr}:AC{adr}']
        for line in range_cell:
            for cell in line:
                cell.fill = PatternFill(fill_type='solid', fgColor="DDDDDD")

        #  Настройка вывода на печать
        self._worksheet = global_print_setting(self._worksheet)
        self._worksheet.page_setup.fitToHeight = 1


class GroupTimesheetRender:
    __style_border = Side(border_style="thin", color="000000")
    _bottom_border = Border(bottom=__style_border)
    _circle_border = Border(top=__style_border, bottom=__style_border, left=__style_border, right=__style_border)

    def __init__(self, month: Month):
        self.month = month
        self._book: Workbook = Workbook()
        self._worksheet: Worksheet = self._book.active
        self.__set_default_value_cells()
        self.__link_title = self.__add_title_sheet()

    @property
    def link_title(self):
        return self.__link_title

    def save_file(self, file_name='test'):
        self.__print_setting()
        try:
            name = f'{file_name}.xlsx'
            self._book.save(name)
        except PermissionError:
            name = check_path_new_file(f'{file_name}.xlsx')
            self._book.save(name)
        return name

    def __print_setting(self):
        self._worksheet = global_print_setting(self._worksheet, left=1, right=1, top=1.05, header=0)
        self._worksheet.page_setup.fitToHeight = False
        self._worksheet.oddFooter.center.text = "Страница &[Page] из &N"  # нижний колонтитул

    def __set_default_value_cells(self):
        # задать ширину ячеек
        for col in get_column_interval(1, 37):
            self._worksheet.column_dimensions[col].width = 3.86
        # задать высоту строк
        for row in range(1, 37):
            self._worksheet.row_dimensions[row].height = 15
        # задать выравниевание в ячейках
        for row in self._worksheet.iter_rows(min_row=1, max_row=36, min_col=1, max_col=36):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                cell.font = Font(size='11')

    def __crate_header(self):
        for i, addr in enumerate(['G1', 'U1', 'W3', 'O4']):
            self._worksheet[addr].value = TIMESHEET_HEADER[i]
            self._worksheet[addr].alignment = Alignment(horizontal='right', vertical='center')
        self._worksheet['W3'].font = Font(size='14')
        self._worksheet['S4'].alignment = Alignment(horizontal='left', vertical='center')
        # self._worksheet['S4'].value = f'{datetime.datetime.now().year} года'

        for range_cell in ['H1:M1', 'V1:AD1', 'P4:R4']:
            self._worksheet.merge_cells(range_cell)
            for line in self._worksheet[range_cell]:
                for cell in line:
                    cell.border = self._bottom_border

    def __create_table_header(self):
        self._worksheet.print_title_rows = '5:8' # задание сквозных строк на каждом листе
        for range_cell in ['A5:B7', 'C5:I7', 'J5:L7', 'M5:AB5', 'AC5:AF5', 'AC6:AD6', 'AE6:AF6',
                           'AC7:AF7', 'AG5:AH6', 'AI5:AJ6', 'AG7:AJ7', 'A8:B8', 'C8:I8', 'J8:L8',
                           'M8:AB8', 'AC8:AD8', 'AE8:AF8', 'AG8:AH8', 'AI8:AJ8']:
            self._worksheet.merge_cells(range_cell)
        for line in self._worksheet['A5:AJ8']:
            for cell in line:
                cell.border = self._circle_border
        for i, addr in enumerate(['A5', 'C5', 'J5', 'M5', 'AC5', 'AC6', 'AE6', 'AG5', 'AI5']):
            self._worksheet[addr].value = TIMESHEET_TAB_HEADER[i]
        self._worksheet['AC7'].value = self._worksheet['AG7'].value = 'дни/часы'
        for val in range(1, self.month.days + 1):
            if val <= self.month.get_means():
                cell = self._worksheet.cell(6, 12 + val)
            else:
                cell = self._worksheet.cell(7, 12 + val - self.month.get_means())
                cell.fill = PatternFill(fill_type='solid', fgColor="DDDDDD")
            cell.value = val
            cell.font = Font(size='9')
        for i, addr in enumerate(['A8', 'C8', 'J8', 'M8', 'AC8', 'AE8', 'AG8', 'AI8']):
            cell = self._worksheet[addr]
            cell.value = i + 1
            cell.font = Font(size='9')

    def __create_footer(self):
        for range_cell in ['L31:P31', 'L34:P34', 'R31:T31', 'V31:Y31', 'R34:T34', 'V34:Y34']:
            for row in self._worksheet[range_cell]:
                for cell in row:
                    cell.border = self._bottom_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        sett: ProgramSetting = ProgramSetting.get_setting()
        for i, addr in enumerate(['K31', 'N31', 'V31', 'K34', 'N34', 'V34']):
            cell = self._worksheet[addr]
            cell.value = TIMESHEET_FOOTER[i]
            if addr.startswith('K'):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif addr.startswith('V'):
                cell.alignment = Alignment(horizontal='left', vertical='center')
        self._worksheet['N31'].value = sett.resp_post
        self._worksheet['N34'].value = sett.head_post
        self._worksheet['V31'].value = sett.resp_name
        self._worksheet['V34'].value = sett.head_name
        for i, addr in enumerate(['N32', 'S32', 'V32', 'AA31', 'N35', 'S35', 'V35', 'AA34']):
            cell = self._worksheet[addr]
            if addr.startswith('AA'):
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.font = Font(vertAlign='superscript', size='11')
                if addr.startswith('V'):
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            if i > 3:
                cell.value = REPEATING_LINES[i - 4]
            else:
                cell.value = REPEATING_LINES[i]

    def __add_title_sheet(self):
        self.__crate_header()
        self.__create_table_header()
        add_rows = [self.__create_row(_) for _ in range(2, 7)]
        self.__create_footer()
        return add_rows

    def __create_row(self, num_row):
        start_row = num_row * 4 + 1
        end_row = num_row * 4 + 4
        for i_line, line in enumerate(
                self._worksheet.iter_rows(min_row=start_row,
                                          max_row=end_row,
                                          min_col=1, max_col=36)):
            for i_col, cell in enumerate(line):
                cell.border = self._circle_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                if i_line > 1 and 12 <= i_col < 28:
                    cell.fill = PatternFill(fill_type='solid', fgColor="DDDDDD")
        for col in [(1, 2), (3, 9), (10, 12), (29, 30), (31, 32), (33, 34), (35, 36)]:
            start = start_row
            end = end_row
            if col[0] > 30:
                self._worksheet.merge_cells(
                    start_row=start, end_row=end - 2, start_column=col[0], end_column=col[1])
                self._worksheet.merge_cells(
                    start_row=start + 2, end_row=end, start_column=col[0], end_column=col[1])
            elif 28 < col[0] < 31:
                for i in range(4):
                    self._worksheet.merge_cells(
                        start_row=start + i, end_row=start + i, start_column=col[0], end_column=col[1])
            else:
                self._worksheet.merge_cells(
                    start_row=start, end_row=end, start_column=col[0], end_column=col[1])

        for col in range(13, 29):
            self._worksheet.merge_cells(
                start_row=start_row, end_row=start_row + 1, start_column=col, end_column=col)
            self._worksheet.merge_cells(
                start_row=end_row - 1, end_row=end_row, start_column=col, end_column=col)
        return start_row

    def add_other_sheet(self, num_sheet):
        offset = (num_sheet - 1) * 8
        add_rows = [self.__create_row(_) for _ in range(9 + offset, 15 + offset)]
        self._worksheet.move_range('A29:AA35', rows=32 * num_sheet)
        self.__create_footer()

        return add_rows

    def fill(self, row, column, value):
        self._worksheet.cell(row, column, value)


class GroupKPIRender:
    # Настройки границ
    __style_border_thin = Side(border_style="thin", color="000000")
    _bottom_border = Border(bottom=__style_border_thin)
    _circle_border = Border(top=__style_border_thin, bottom=__style_border_thin,
                            left=__style_border_thin, right=__style_border_thin)

    # Стили шрифтов
    _font_bold = Font(bold=True, size=11)
    _font_small = Font(size=8)
    _header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    def __init__(self, month: Month):
        self.sett: ProgramSetting = ProgramSetting.get_setting()
        self.month = month
        self._book = Workbook()
        self._worksheet = self._book.active
        self._worksheet.title = f"KPI {self.month.number}-{self.month.year}"

        # Настройка мелкой сетки (как в вашем TimeSheet)
        for col in range(1, 41):
            self._worksheet.column_dimensions[get_column_letter(col)].width = 3.0

        self.__create_header()

    def __print_setting(self):
        self._worksheet = global_print_setting(self._worksheet, left=1, right=1, top=1.05, header=0)
        self._worksheet.page_setup.fitToHeight = False
        self._worksheet.oddFooter.center.text = "Страница &[Page] из &N"  # нижний колонтитул

    def __create_header(self):
        ws = self._worksheet

        # 1. Организация и Подразделение
        # Строка 1: Организация
        ws.merge_cells('B1:F1')
        ws['B1'] = "Организация:"
        ws['B1'].alignment = Alignment(horizontal='right')

        ws.merge_cells('G1:O1')
        ws['G1'] = self.sett.org
        self.__apply_border_range(1, 7, 15, border_type='bottom')
        ws['G1'].alignment = Alignment(horizontal='center')

        # Строка 1: Подразделение
        ws.merge_cells('Q1:V1')
        ws['Q1'] = "Подразделение:"
        ws['Q1'].alignment = Alignment(horizontal='right')

        ws.merge_cells('W1:AJ1')
        ws['W1'] = self.sett.div
        self.__apply_border_range(1, 23, 36, border_type='bottom')
        ws['W1'].alignment = Alignment(horizontal='center')

        # 2. Заголовок документа
        ws.merge_cells('A3:AJ3')
        title = ws['A3']
        title.value = "ОТЧЕТ ПО ПОКАЗАТЕЛЯМ KPI"
        title.font = Font(bold=True, size=14)
        title.alignment = Alignment(horizontal='center')

        ws.merge_cells('A4:AJ4')
        subtitle = ws['A4']
        subtitle.value = f"за {self.month.get_lower()} {self.month.year} года"
        subtitle.alignment = Alignment(horizontal='center')

        # 3. Шапка таблицы (строка 6)
        mapping = [
            ('A6:B6', '№'),
            ('C6:O6', 'Сотрудник'),
            ('P6:V6', 'План'),
            ('W6:AC6', 'Факт'),
            ('AD6:AJ6', 'KPI')
        ]
        for cells, text in mapping:
            ws.merge_cells(cells)
            cell = ws[cells.split(':')[0]]
            cell.value = text
            cell.font = self._font_bold
            cell.alignment = Alignment(horizontal='center')
            cell.fill = self._header_fill
            self.__apply_border_range_by_str(cells, border_type='all')

    def fill_data(self, query):
        ws = self._worksheet
        start_row = 7
        last_row = start_row

        for i, worker in enumerate(query):
            last_row = start_row + i

            # Расчет KPI
            kpi = (worker.total_plan / worker.total_fact) if worker.total_fact else 0

            # Данные строки и их соответствие столбцам (начало:конец колонки)
            row_map = [
                (1, 2, i + 1),  # №
                (3, 15, worker.get_short_name()),  # Сотрудник
                (16, 22, worker.total_plan),  # План
                (23, 29, worker.total_fact),  # Факт
                (30, 36, round(kpi, 2))  # KPI
            ]

            for c_start, c_end, val in row_map:
                ws.merge_cells(start_row=last_row, end_row=last_row, start_column=c_start, end_column=c_end)
                cell = ws.cell(row=last_row, column=c_start, value=val)
                self.__apply_border_range(last_row, c_start, c_end, border_type='all')

                if c_start == 3:  # ФИО
                    cell.alignment = Alignment(horizontal='left', indent=1)
                else:
                    cell.alignment = Alignment(horizontal='center')

        self.__create_total_row(start_row, last_row)
        self.__create_footer(last_row + 2)

    def __create_total_row(self, start_row, last_row):
        ws = self._worksheet
        tr = last_row + 1

        # Итоговая подпись (A:O)
        ws.merge_cells(start_row=tr, end_row=tr, start_column=1, end_column=15)
        ws.cell(row=tr, column=1).value = "ИТОГО ПО ПОДРАЗДЕЛЕНИЮ:"
        ws.cell(row=tr, column=1).font = self._font_bold
        ws.cell(row=tr, column=1).alignment = Alignment(horizontal='right')

        # Суммы для Плана (P:V) и Факта (W:AC)
        # Используем column_index_from_string для безопасного получения номера колонки
        sums = [('P', 'V'), ('W', 'AC')]
        for s_col, e_col in sums:
            start_c = column_index_from_string(s_col)
            end_c = column_index_from_string(e_col)

            ws.merge_cells(start_row=tr, end_row=tr, start_column=start_c, end_column=end_c)
            cell = ws.cell(row=tr, column=start_c)
            # Формула SUM(P7:P20)
            cell.value = f"=SUM({s_col}{start_row}:{s_col}{last_row})"
            cell.font = self._font_bold
            self.__apply_border_range(tr, start_c, end_c, border_type='all')

        # Итоговый KPI (AD:AJ)
        start_kpi = column_index_from_string('AD')
        end_kpi = column_index_from_string('AJ')
        ws.merge_cells(start_row=tr, end_row=tr, start_column=start_kpi, end_column=end_kpi)

        cell_kpi = ws.cell(row=tr, column=start_kpi)
        # Формула План/Факт: P_итого / W_итого
        cell_kpi.value = f"=IF(W{tr}>0, P{tr}/W{tr}, 0)"
        cell_kpi.font = self._font_bold
        cell_kpi.number_format = '0.00'
        self.__apply_border_range(tr, start_kpi, end_kpi, border_type='all')

    def __create_footer(self, r):
        r += 1
        ws = self._worksheet
        signatures = [
            ("Ответственное лицо:", self.sett.resp_post, self.sett.resp_name),
            ("Руководитель подразделения:", self.sett.head_post, self.sett.head_name)
        ]

        for label, pos, name in signatures:
            # Метка
            ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=12)
            ws.cell(r, 1, label).alignment = Alignment(horizontal='right')

            # Должность
            ws.merge_cells(start_row=r, end_row=r, start_column=14, end_column=20)
            ws.cell(r, 14, pos).alignment = Alignment(horizontal='center')
            self.__apply_border_range(r, 14, 20, border_type='bottom')

            ws.merge_cells(start_row=r + 1, end_row=r + 1, start_column=14, end_column=20)
            sub = ws.cell(r + 1, 14, "должность")
            sub.font = self._font_small
            sub.alignment = Alignment(horizontal='center', vertical='top')

            # Подпись
            ws.merge_cells(start_row=r, end_row=r, start_column=22, end_column=26)
            self.__apply_border_range(r, 22, 26, border_type='bottom')

            ws.merge_cells(start_row=r + 1, end_row=r + 1, start_column=22, end_column=26)
            sub_p = ws.cell(r + 1, 22, "подпись")
            sub_p.font = self._font_small
            sub_p.alignment = Alignment(horizontal='center', vertical='top')

            # Расшифровка
            ws.merge_cells(start_row=r, end_row=r, start_column=28, end_column=34)
            ws.cell(r, 28, name).alignment = Alignment(horizontal='center')
            self.__apply_border_range(r, 28, 34, border_type='bottom')

            ws.merge_cells(start_row=r + 1, end_row=r + 1, start_column=28, end_column=34)
            sub_r = ws.cell(r + 1, 28, "расшифровка подписи")
            sub_r.font = self._font_small
            sub_r.alignment = Alignment(horizontal='center', vertical='top')

            # Дата
            ws.merge_cells(start_row=r, end_row=r, start_column=36, end_column=40)
            ws.cell(r, 36, f'"  " _________ 20{str(self.month.year)[2:]} г.').alignment = Alignment(horizontal='left')

            r += 3

    # Вспомогательные методы для границ
    def __apply_border_range(self, row, c_start, c_end, border_type='all'):
        border = self._circle_border if border_type == 'all' else self._bottom_border
        for col in range(c_start, c_end + 1):
            self._worksheet.cell(row, col).border = border

    def __apply_border_range_by_str(self, range_str, border_type='all'):
        border = self._circle_border if border_type == 'all' else self._bottom_border
        for row in self._worksheet[range_str]:
            for cell in row:
                cell.border = border

    def save_file(self):
        self.__print_setting()
        try:
            filename = f'kpi-{self.month.number}-{self.month.year}.xlsx'
            self._book.save(filename)
        except PermissionError:
            filename = check_path_new_file(f'kpi-{self.month.number}-{self.month.year}.xlsx')
            self._book.save(filename)

        return filename
