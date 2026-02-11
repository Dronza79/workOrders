import datetime

from openpyxl import Workbook
from openpyxl.styles import Side, Border, Alignment, Font, PatternFill
from openpyxl.utils import get_column_interval
from openpyxl.worksheet.worksheet import Worksheet

from database.models import Month
from .template_text import TEMPLATE_ALIGNMENT_RIGHT, TEMPLATE_CLARIFICATION, TEMPLATE_TABLE, TIMESHEET_HEADER, \
    TIMESHEET_TAB_HEADER, TIMESHEET_FOOTER, REPEATING_LINES
from .utils import global_print_setting, check_path_new_file


# from tablesExcel.forms import TimeSheet


class PersonalMonthExelTable:
    def __init__(self):
        self._book: Workbook = Workbook()
        self._worksheet: Worksheet = self._book.active
        self._worksheet.title = '1'
        self.prefill_worksheet()

    def save(self, file_name='test'):
        try:
            name = f'{file_name}.xlsx'
            self._book.save(name)
        except PermissionError:
            name = check_path_new_file(f'{file_name}.xlsx')
            self._book.save(name)
        return name

    def add_worksheet(self):
        self._worksheet = self._book.create_sheet(f'{len(self._book.sheetnames) + 1}')
        self.prefill_worksheet()

    def fill_data(self, worker, month, calc, target, data_task, data_old_task):
        offset = 0
        for num in range(1, month.days + 1):
            if num < 16:
                cell = self._worksheet.cell(7, 13 + num)
            else:
                cell = self._worksheet.cell(8, num - 2)
            cell.value = num
            cell.font = Font(size='9')
        # print(f'fill_data {worker=} {worker.function=}')

        self._worksheet.cell(1, 20).value = f'{month.year} года'
        self._worksheet.cell(1, 18).value = str(month).lower()
        self._worksheet.cell(4, 28).value = target
        self._worksheet.cell(4, 4).value = worker.get_short_name()
        self._worksheet.cell(4, 12).value = worker.table_num
        self._worksheet.cell(4, 17).value = worker.function.post
        self._worksheet.cell(28, 13).value = worker.get_short_name()
        self._worksheet.cell(21, 30).value = sum(calc.get('first_half', 0))
        self._worksheet.cell(23, 30).value = sum(calc.get('second_half', 0))

        self._worksheet.cell(4, 4).alignment = Alignment(horizontal='center', vertical='center')
        self._worksheet.cell(4, 12).alignment = Alignment(horizontal='center', vertical='center')
        self._worksheet.cell(4, 17).alignment = Alignment(horizontal='center', vertical='center')
        self._worksheet.cell(4, 28).alignment = Alignment(horizontal='center', vertical='center')
        for i, task in enumerate(data_task):
            self._worksheet.cell(9 + offset + i, 1).value = i + 1
            data_string = f'{task.is_type} '
            data_string += f'{task.order}\n' if task.order else '\n'
            data_string += f'{task.order.type_obj} {task.order.title} {task.order.name}\n' if task.order else ''
            data_string += f'{task.order.article}' if task.order else f'{task.comment}'
            self._worksheet.cell(9 + offset + i, 2).value = data_string
            self._worksheet.cell(9 + offset + i, 2).alignment += Alignment(wrapText=True)
            self._worksheet.cell(9 + offset + i, 12).value = task.deadline - sum(data_old_task[i].time_worked)
            self._worksheet.cell(9 + offset + i, 32).value = sum(data_old_task[i].time_worked)
            self._worksheet.cell(9 + offset + i, 30).value = sum(task.time_worked)
            for period in task.time_worked:
                if period.date.day <= 15:
                    self._worksheet.cell(9 + offset + i, 13 + period.date.day).value = period.value
                else:
                    self._worksheet.cell(9 + offset + i + 1, period.date.day - 2).value = period.value
            offset += 1

    def prefill_worksheet(self):

        # задать высоту строк
        row_height = [15 for _ in range(5)]
        row_height += [18, 16.5, 16.5]
        row_height += [29.5 for _ in range(12)]
        row_height += [15 for _ in range(10)]
        for key, val in enumerate(row_height):
            # self._worksheet.row_dimensions[key + 6].height = val  # реальное значение
            self._worksheet.row_dimensions[key + 1].height = val  # реальное значение

        # задать ширину ячеек
        for idx in get_column_interval(2, 31):
            self._worksheet.column_dimensions[idx].width = 5  # значение в пикселях умноженное на 7
        self._worksheet.column_dimensions['A'].width = 6

        # объединение ячеек
        lst_range_cell = ['A6:A8', 'B6:K8', 'L6:M8', 'N6:AC6', 'AD6:AE8']
        lst_range_cell += ['A9:A10', 'B9:K10', 'L9:M10', 'AD9:AE10']
        lst_range_cell += ['A11:A12', 'B11:K12', 'L11:M12', 'AD11:AE12']
        lst_range_cell += ['A13:A14', 'B13:K14', 'L13:M14', 'AD13:AE14']
        lst_range_cell += ['A15:A16', 'B15:K16', 'L15:M16', 'AD15:AE16']
        lst_range_cell += ['A17:A18', 'B17:K18', 'L17:M18', 'AD17:AE18']
        lst_range_cell += ['A19:A20', 'B19:K20', 'L19:M20', 'AD19:AE20']
        lst_range_cell += ['W21:AC22', 'W23:AC24', 'AD21:AE22', 'AD23:AE24', 'AD25:AE26', 'AB25:AC26']
        lst_range_cell += ['D4:G4', 'l4:M4', 'Q4:U4', 'AB4:AC4', 'R1:S1']
        [self._worksheet.merge_cells(range_string) for range_string in lst_range_cell]

        # границы и выравнивание в ячейках таблицы
        table = self._worksheet['A6:AE20'] + self._worksheet['AD21:AE26']
        thins = Side(border_style="thin", color="000000")
        for line in table:
            for cell in line:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)

        # нижнее подчеркивание
        underline = (self._worksheet['D4:G4'] + self._worksheet['L4:M4'] + self._worksheet['Q4:U4'] +
                     self._worksheet['R1:S1'] + self._worksheet['H22:M22'] + self._worksheet['H25:M25'] +
                     self._worksheet['H28:M28'] + self._worksheet['AB4:AC4'])
        for line in underline:
            for cell in line:
                cell.border = Border(bottom=thins)

        # предзаполнение текста заголовка и подписей экселя
        for i, adr in enumerate(['Q1', 'K4', 'P4', 'AA4', 'G22', 'M22', 'G25', 'M25', 'M28']):
            cell = self._worksheet[adr]
            if adr in ['M22', 'M25', 'M28']:
                cell.font = Font(size='11')
            else:
                cell.font = Font(size='14')
            if adr != 'M28':
                cell.value = TEMPLATE_ALIGNMENT_RIGHT[i]
            cell.alignment = Alignment(horizontal='right', vertical='center')
        for adr in ['C4', 'G28']:
            cell = self._worksheet[adr]
            cell.font = Font(size='14')
            cell.value = TEMPLATE_ALIGNMENT_RIGHT[-1]
            cell.alignment = Alignment(horizontal='right', vertical='center')
        self._worksheet['T1'].alignment = Alignment(horizontal='left', vertical='center')
        self._worksheet['T1'].font = Font(size='14')

        # заполнение надстрочных пояснений
        for adr in ['D5', 'S5', 'H23', 'H26', 'H29']:
            self._worksheet[adr].font = Font(vertAlign='superscript', size='14')
            if adr != 'S5':
                self._worksheet[adr].alignment = Alignment(horizontal='left', vertical='center')
                if adr == 'D5':
                    self._worksheet[adr].value = TEMPLATE_CLARIFICATION[0]
                else:
                    self._worksheet[adr].value = TEMPLATE_CLARIFICATION[-1]
            else:
                self._worksheet[adr].alignment = Alignment(horizontal='center', vertical='center')
                self._worksheet[adr].value = TEMPLATE_CLARIFICATION[1]

        # предзаполнение таблицы
        for i, adr in enumerate(['A6', 'B6', 'L6', 'N6', 'AD6', 'W21', 'W23', 'AB25']):
            self._worksheet[adr].value = TEMPLATE_TABLE[i]
            self._worksheet[adr].font = Font(size='11')
            self._worksheet[adr].alignment += Alignment(wrapText=True)
            if adr in ['W21', 'W23', 'AB25']:
                self._worksheet[adr].alignment += Alignment(horizontal='right', vertical='center')
        self._worksheet['AD25'] = '=SUM(AD21:AE24)'

        # заливка ячеек таблицы
        range_cell = []
        for adr in range(8, 21, 2):
            range_cell += self._worksheet[f'N{adr}:AC{adr}']
        for line in range_cell:
            for cell in line:
                cell.fill = PatternFill(fill_type='solid', fgColor="DDDDDD")

        #  Настройка вывода на печать
        self._worksheet = global_print_setting(self._worksheet)
        self._worksheet.page_setup.fitToHeight = 1


class TimeSheet:
    __style_border = Side(border_style="thin", color="000000")
    _bottom_border = Border(bottom=__style_border)
    _circle_border = Border(top=__style_border, bottom=__style_border, left=__style_border, right=__style_border)

    def __init__(self, month: Month):
        self.month = month
        self._book: Workbook = Workbook()
        self._worksheet: Worksheet = self._book.active
        self.__set_default_value_cells()
        self.__link_title = self.__add_title_sheet()

    @property
    def link_title(self):
        return self.__link_title

    def save(self, file_name='test'):
        self.__print_setting()
        try:
            name = f'{file_name}.xlsx'
            self._book.save(name)
        except PermissionError:
            name = check_path_new_file(f'{file_name}.xlsx')
            self._book.save(name)
        return name

    def __print_setting(self):
        self._worksheet = global_print_setting(self._worksheet, left=1, right=1, top=1.05, header=0)
        self._worksheet.page_setup.fitToHeight = False
        self._worksheet.oddFooter.center.text = "Страница &[Page] из &N"  # нижний колонтитул

    def __set_default_value_cells(self):
        # задать ширину ячеек
        for col in get_column_interval(1, 37):
            self._worksheet.column_dimensions[col].width = 3.86
        # задать высоту строк
        for row in range(1, 37):
            self._worksheet.row_dimensions[row].height = 15
        # задать выравниевание в ячейках
        for row in self._worksheet.iter_rows(min_row=1, max_row=36, min_col=1, max_col=36):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                cell.font = Font(size='11')

    def __crate_header(self):
        for i, addr in enumerate(['G1', 'U1', 'W3', 'O4']):
            self._worksheet[addr].value = TIMESHEET_HEADER[i]
            self._worksheet[addr].alignment = Alignment(horizontal='right', vertical='center')
        self._worksheet['W3'].font = Font(size='14')
        self._worksheet['S4'].alignment = Alignment(horizontal='left', vertical='center')
        # self._worksheet['S4'].value = f'{datetime.datetime.now().year} года'

        for range_cell in ['H1:M1', 'V1:AD1', 'P4:R4']:
            self._worksheet.merge_cells(range_cell)
            for line in self._worksheet[range_cell]:
                for cell in line:
                    cell.border = self._bottom_border

    def __create_table_header(self):
        self._worksheet.print_title_rows = '5:8' # задание сквозных строк на каждом листе
        for range_cell in ['A5:B7', 'C5:I7', 'J5:L7', 'M5:AB5', 'AC5:AF5', 'AC6:AD6', 'AE6:AF6',
                           'AC7:AF7', 'AG5:AH6', 'AI5:AJ6', 'AG7:AJ7', 'A8:B8', 'C8:I8', 'J8:L8',
                           'M8:AB8', 'AC8:AD8', 'AE8:AF8', 'AG8:AH8', 'AI8:AJ8']:
            self._worksheet.merge_cells(range_cell)
        for line in self._worksheet['A5:AJ8']:
            for cell in line:
                cell.border = self._circle_border
        for i, addr in enumerate(['A5', 'C5', 'J5', 'M5', 'AC5', 'AC6', 'AE6', 'AG5', 'AI5']):
            self._worksheet[addr].value = TIMESHEET_TAB_HEADER[i]
        self._worksheet['AC7'].value = self._worksheet['AG7'].value = 'дни/часы'
        for val in range(1, self.month.days + 1):
            print(f'{val=}')
            if val <= self.month.get_means():
                cell = self._worksheet.cell(6, 12 + val)
            else:
                cell = self._worksheet.cell(7, 12 + val - self.month.get_means())
                cell.fill = PatternFill(fill_type='solid', fgColor="DDDDDD")
            print(f'{cell=}')
            cell.value = val
            cell.font = Font(size='9')
        for i, addr in enumerate(['A8', 'C8', 'J8', 'M8', 'AC8', 'AE8', 'AG8', 'AI8']):
            cell = self._worksheet[addr]
            cell.value = i + 1
            cell.font = Font(size='9')

    def __create_footer(self):
        for range_cell in ['L31:P31', 'L34:P34', 'R31:T31', 'V31:Y31', 'R34:T34', 'V34:Y34']:
            for row in self._worksheet[range_cell]:
                for cell in row:
                    cell.border = self._bottom_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        for i, addr in enumerate(['K31', 'N31', 'V31', 'K34', 'N34', 'V34']):
            cell = self._worksheet[addr]
            cell.value = TIMESHEET_FOOTER[i]
            if addr.startswith('K'):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif addr.startswith('V'):
                cell.alignment = Alignment(horizontal='left', vertical='center')
        for i, addr in enumerate(['N32', 'S32', 'V32', 'AA31', 'N35', 'S35', 'V35', 'AA34']):
            cell = self._worksheet[addr]
            if addr.startswith('AA'):
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.font = Font(vertAlign='superscript', size='11')
                if addr.startswith('V'):
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            if i > 3:
                cell.value = REPEATING_LINES[i - 4]
            else:
                cell.value = REPEATING_LINES[i]

    def __add_title_sheet(self):
        self.__crate_header()
        self.__create_table_header()
        add_rows = [self.__create_row(_) for _ in range(2, 7)]
        self.__create_footer()
        return add_rows

    def __create_row(self, num_row):
        start_row = num_row * 4 + 1
        end_row = num_row * 4 + 4
        # print(f'{start_row=} {end_row=}')
        for i_line, line in enumerate(
                self._worksheet.iter_rows(min_row=start_row,
                                          max_row=end_row,
                                          min_col=1, max_col=36)):
            for i_col, cell in enumerate(line):
                cell.border = self._circle_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                if i_line > 1 and 12 <= i_col < 28:
                    cell.fill = PatternFill(fill_type='solid', fgColor="DDDDDD")
        for col in [(1, 2), (3, 9), (10, 12), (29, 30), (31, 32), (33, 34), (35, 36)]:
            start = start_row
            end = end_row
            if col[0] > 30:
                self._worksheet.merge_cells(
                    start_row=start, end_row=end - 2, start_column=col[0], end_column=col[1])
                self._worksheet.merge_cells(
                    start_row=start + 2, end_row=end, start_column=col[0], end_column=col[1])
            elif 28 < col[0] < 31:
                for i in range(4):
                    self._worksheet.merge_cells(
                        start_row=start + i, end_row=start + i, start_column=col[0], end_column=col[1])
            else:
                self._worksheet.merge_cells(
                    start_row=start, end_row=end, start_column=col[0], end_column=col[1])

        for col in range(13, 29):
            self._worksheet.merge_cells(
                start_row=start_row, end_row=start_row + 1, start_column=col, end_column=col)
            self._worksheet.merge_cells(
                start_row=end_row - 1, end_row=end_row, start_column=col, end_column=col)
        return start_row

    def add_other_sheet(self, num_sheet):
        offset = (num_sheet - 1) * 8
        add_rows = [self.__create_row(_) for _ in range(9 + offset, 15 + offset)]
        self._worksheet.move_range('A29:AA35', rows=32 * num_sheet)
        self.__create_footer()

        return add_rows

    def fill(self, row, column, value):
        self._worksheet.cell(row, column, value)

